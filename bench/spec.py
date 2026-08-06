"""NIST のテストケース定義（xlsx）を読んで、抽出器を外から検算する。

この xlsx は各ファイルの全PMIの一覧ではない。「この部品がどのテスト項目(ATC)を
実証しているか」のチェックリストである（51行）。したがって全数検算はできない。
挙がっている項目についてだけ、外部の権威と突き合わせられる。

**読めなかった行は黙って捨てない。** 捨てると「検算した」ことにならないので、
必ず未解釈として数え上げて報告する。
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "corpus" / "nist" / "NIST-FTC-PMI-Definitions.xlsx"

# 幾何特性記号 -> STEP の実体名
SYMBOL_KIND = {
    "⌖": "POSITION_TOLERANCE",  # ⌖
    "⌓": "SURFACE_PROFILE_TOLERANCE",  # ⌓
    "⌢": "LINE_PROFILE_TOLERANCE",  # ⌢
    "⫽": "PARALLELISM_TOLERANCE",  # ⫽
    "⏊": "PERPENDICULARITY_TOLERANCE",  # ⏊
    "▱": "FLATNESS_TOLERANCE",  # ▱
    "⌭": "CYLINDRICITY_TOLERANCE",  # ⌭
    "⌯": "SYMMETRY_TOLERANCE",  # ⌯
}

# 材料状態などの修飾子
SYMBOL_MODIFIER = {
    "Ⓜ": "MAXIMUM_MATERIAL_REQUIREMENT",  # Ⓜ
    "Ⓛ": "LEAST_MATERIAL_REQUIREMENT",  # Ⓛ
    "Ⓟ": "PROJECTED_TOLERANCE_ZONE",  # Ⓟ
    "Ⓕ": "FREE_STATE",  # Ⓕ
    "Ⓣ": "TANGENT_PLANE",  # Ⓣ
}

DIAMETER = "⌀Ø"  # ⌀ と Ø の両方が使われている
_NUM = re.compile(r"(\d*\.\d+|\d+\.?\d*)")
# 共通データムは2つ組とは限らない。FTC-10 に E-F-G の3つ組が出る。
_DATUM = re.compile(r"^([A-Z](?:-[A-Z])*)$")


@dataclass
class Claim:
    """定義から読み取った「このPMIが入っているはず」1件。"""

    ftc: str
    atc: str
    category: str
    kind: str
    value: float | None
    zone_form: str  # 'spherical' / 'cylindrical or circular' / ''
    datums: tuple[str, ...]
    modifiers: tuple[str, ...]
    raw: str


@dataclass
class SpecTable:
    claims: list[Claim] = field(default_factory=list)
    unparsed: list[tuple[str, str, str]] = field(default_factory=list)  # (ftc, atc, raw)
    rows: int = 0


def _parse_fcf(text: str) -> list[dict]:
    """公差記入枠を読む。1つのセルに複数入ることがある。

        ⌖ | S⌀ .025 | D | B | C
        ⌖ | Ø.040Ⓜ | ⌀.045 MAX | A
        ⫽ | .03 | A
    """
    out: list[dict] = []
    for line in re.split(r"[\n\r]+", text):
        for sym, kind in SYMBOL_KIND.items():
            idx = line.find(sym)
            if idx < 0:
                continue
            body = line[idx + 1 :]
            segs = [s.strip() for s in body.split("|")]
            segs = [s for s in segs if s]
            if not segs:
                continue

            head = segs[0]
            # 「.01 / Ø1.00」の / 以降は単位面積あたりの限定であって公差域の形ではない
            zone_src = head.split("/")[0]
            zone = ""
            if re.search(r"S\s*[⌀Ø]", zone_src):
                zone = "spherical"
            elif any(c in zone_src for c in DIAMETER):
                zone = "cylindrical or circular"
            m = _NUM.search(head)
            value = float(m.group(1)) if m else None
            mods = tuple(SYMBOL_MODIFIER[c] for c in head if c in SYMBOL_MODIFIER)

            datums: list[str] = []
            for s in segs[1:]:
                token = "".join(ch for ch in s if ch not in SYMBOL_MODIFIER).strip()
                token = token.replace(" ", "")
                if _DATUM.match(token):
                    datums.append(token)
            out.append(
                {"kind": kind, "value": value, "zone_form": zone,
                 "datums": tuple(datums), "modifiers": mods, "raw": line.strip()}
            )
            break
    return out


def load(path: Path = XLSX) -> SpecTable:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    table = SpecTable()
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[0]:
            continue
        ftc = str(r[0]).strip()
        if not ftc.isdigit():  # 途中に挟まる見出し行
            continue
        table.rows += 1
        atc = str(r[1]).strip() if r[1] else ""
        category = str(r[2]).strip() if r[2] else ""
        raw = str(r[4]).strip() if r[4] else ""
        found = _parse_fcf(raw)
        if not found:
            # 幾何公差の記号が無い行（寸法・データムターゲット等）は検算対象外。
            # 記号があるのに読めなかったものだけを未解釈として数える。
            if any(sym in raw for sym in SYMBOL_KIND):
                table.unparsed.append((ftc, atc, raw))
            continue
        for f in found:
            table.claims.append(
                Claim(ftc=ftc, atc=atc, category=category, kind=f["kind"], value=f["value"],
                      zone_form=f["zone_form"], datums=f["datums"],
                      modifiers=f["modifiers"], raw=f["raw"])
            )
    return table
